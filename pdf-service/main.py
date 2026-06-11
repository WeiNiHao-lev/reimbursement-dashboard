import copy
import os
import shutil
import subprocess
import tempfile
import json
import traceback
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from fastapi import FastAPI, File, Form, UploadFile, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from PyPDF2 import PdfMerger
from PIL import Image
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Image as RLImage

BASE = Path(__file__).parent
COVER_TEMPLATE   = BASE / "CCEPC Template Reimbursement (Cover).xlsx"
SUMMARY_TEMPLATE = BASE / "CCEPC Template Reimbursement (Summary for Page 2).xlsx"

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])


# ── LibreOffice ───────────────────────────────────────────────────────────────

def find_soffice() -> Optional[str]:
    for path in ["/usr/bin/soffice", "/usr/lib/libreoffice/program/soffice", shutil.which("soffice")]:
        if path and Path(path).exists():
            return path
    return None


def excel_to_pdf(xlsx_path: Path, out_dir: Path) -> Path:
    soffice = find_soffice()
    if not soffice:
        raise RuntimeError("LibreOffice not found")
    subprocess.run(
        [soffice, "--headless", "--convert-to", "pdf", "--outdir", str(out_dir), str(xlsx_path)],
        check=True, capture_output=True, timeout=60,
        env={**os.environ, "HOME": str(out_dir)},
    )
    return out_dir / (xlsx_path.stem + ".pdf")


# ── openpyxl helpers ──────────────────────────────────────────────────────────

def _copy_row_style(ws, src_row: int, dst_row: int):
    """Copy styles + row height from src_row to dst_row (skip merge slaves)."""
    for col_idx in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col_idx)
        dst = ws.cell(dst_row, col_idx)
        if isinstance(src, MergedCell) or isinstance(dst, MergedCell):
            continue
        if src.has_style:
            dst.font       = copy.copy(src.font)
            dst.border     = copy.copy(src.border)
            dst.fill       = copy.copy(src.fill)
            dst.number_format = src.number_format
            dst.protection = copy.copy(src.protection)
            dst.alignment  = copy.copy(src.alignment)
    if src_row in ws.row_dimensions:
        ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height


def _insert_rows_safe(ws, at: int, count: int, style_src: int):
    """
    Insert `count` rows at `at`, then copy style from `style_src`.

    Merges that CROSS the insertion boundary (min_row < at <= max_row) are split:
    - The part above `at` is re-merged as before.
    - The inserted rows are left as individual (writable) cells.
    - The part below (at + count onwards) is re-merged.

    This preserves section-title merges above the insertion without
    accidentally unmerging them.
    """
    # Snapshot merges that will be extended across the insertion
    crossing = [
        (m.min_row, m.max_row, m.min_col, m.max_col)
        for m in list(ws.merged_cells.ranges)
        if m.min_row < at <= m.max_row
    ]

    ws.insert_rows(at, count)

    for r0, r1, c0, c1 in crossing:
        c0l = get_column_letter(c0)
        c1l = get_column_letter(c1)
        new_r1 = r1 + count  # original max row has shifted down

        # Remove the (now extended) merge
        try:
            ws.unmerge_cells(f"{c0l}{r0}:{c1l}{new_r1}")
        except Exception:
            pass

        # Re-merge the portion ABOVE the insertion (needs ≥ 2 rows)
        if at - 1 > r0:
            try:
                ws.merge_cells(f"{c0l}{r0}:{c1l}{at - 1}")
            except Exception:
                pass

        # Re-merge the portion BELOW the insertion (needs ≥ 2 rows)
        below = at + count
        if new_r1 > below:
            try:
                ws.merge_cells(f"{c0l}{below}:{c1l}{new_r1}")
            except Exception:
                pass

    # Copy style from reference row to each newly inserted row
    for i in range(count):
        _copy_row_style(ws, style_src, at + i)


def _wc(ws, row: int, col: str, value, wrap: bool = False):
    """Write value to cell, skip MergedCell slaves."""
    cell = ws[f"{col}{row}"]
    if isinstance(cell, MergedCell):
        return
    cell.value = value
    if wrap and value is not None:
        try:
            al = cell.alignment
            cell.alignment = Alignment(
                wrap_text=True,
                horizontal=al.horizontal if al else "center",
                vertical=al.vertical   if al else "center",
            )
        except Exception:
            pass


# ── Cover page ────────────────────────────────────────────────────────────────

def fill_cover(form: dict, out_path: Path):
    wb = openpyxl.load_workbook(COVER_TEMPLATE)
    is_domestic = form.get("tripType") == "domestic"
    active_name = "差旅费报销明细（境内）" if is_domestic else "差旅费报销明细（境外）"
    hidden_name = "差旅费报销明细（境外）" if is_domestic else "差旅费报销明细（境内）"

    if hidden_name in wb.sheetnames:
        del wb[hidden_name]

    wb.active = wb[active_name]
    ws = wb.active

    # ── Collect data ──────────────────────────────────────────────────────────
    trip_info     = form.get("tripInfo", [])
    accommodation = form.get("accommodation", [])
    receipts      = form.get("receipts", [])

    urban_by_route:     dict = {}
    intercity_by_route: dict = {}
    for r in receipts:
        if r.get("currency") != "IDR":
            continue
        origin = r.get("origin", "").split(",")[0].strip()
        dest   = r.get("destination", "").split(",")[0].strip()
        route  = f"{origin} - {dest}" if (origin or dest) else "Unknown"
        if r.get("category") == "transportation_urban":
            urban_by_route[route] = urban_by_route.get(route, 0) + r.get("amount", 0)
        elif r.get("category") == "transportation_intercity":
            intercity_by_route[route] = intercity_by_route.get(route, 0) + r.get("amount", 0)

    all_transport = list(dict.fromkeys(list(urban_by_route) + list(intercity_by_route)))

    TRIP_R0  = 7;  TRIP_CAP  = 3
    TRNS_R0  = 13; TRNS_CAP  = 8
    ACCOM_R0 = 23; ACCOM_CAP = 3
    ALLOW_R0 = 29

    extra_trip  = max(0, len(trip_info)     - TRIP_CAP)
    extra_trns  = max(0, len(all_transport) - TRNS_CAP)
    extra_accom = max(0, len(accommodation) - ACCOM_CAP)
    total_extra = extra_trip + extra_trns + extra_accom

    # ── Insert extra rows bottom → top using safe merge-splitting ─────────────
    # (bottom first so earlier row numbers stay stable)
    if extra_accom > 0:
        at = ACCOM_R0 + ACCOM_CAP
        _insert_rows_safe(ws, at, extra_accom, at - 1)
    if extra_trns > 0:
        at = TRNS_R0 + TRNS_CAP
        _insert_rows_safe(ws, at, extra_trns, at - 1)
    if extra_trip > 0:
        at = TRIP_R0 + TRIP_CAP
        _insert_rows_safe(ws, at, extra_trip, at - 1)

    # ── Recalculate section starts after insertions ───────────────────────────
    trns_r  = TRNS_R0  + extra_trip
    accom_r = ACCOM_R0 + extra_trip + extra_trns
    allow_r = ALLOW_R0 + extra_trip + extra_trns + extra_accom

    # ── Page setup ────────────────────────────────────────────────────────────
    # Preserve the template's native page setup — it is already configured to
    # produce a correct A4 single-page PDF. Only add horizontal centering and
    # update the print area for any rows we inserted above.
    ws.print_options.horizontalCentered = True

    # ── Set print area ────────────────────────────────────────────────────────
    # Read the template's existing end row and extend it by the number of rows
    # we inserted. ALWAYS hardcode column "D" — the title rows may have merged
    # cells reaching column H or beyond, making max_column >> 4 and doubling
    # the rendered page width when fed to LibreOffice.
    import re as _re
    orig_end = 35  # safe fallback
    orig_pa = ws.print_area  # e.g. "$A$1:$D$35" or "A1:D35"
    if orig_pa:
        m = _re.search(r'\$?[A-Z]+\$?(\d+)\s*$', orig_pa)
        if m:
            orig_end = int(m.group(1))
    ws.print_area = f"A1:D{orig_end + total_extra}"

    # ── Clear old template data ───────────────────────────────────────────────
    for row in range(TRIP_R0, TRIP_R0 + max(TRIP_CAP, len(trip_info))):
        for c in ["A","B","C","D"]: _wc(ws, row, c, None)
    for row in range(trns_r,  trns_r  + max(TRNS_CAP,  len(all_transport))):
        for c in ["A","B","C","D"]: _wc(ws, row, c, None)
    for row in range(accom_r, accom_r + max(ACCOM_CAP, len(accommodation))):
        for c in ["A","B","C"]: _wc(ws, row, c, None)
    for row in range(allow_r, allow_r + 2):
        for c in ["A","B","C","D"]: _wc(ws, row, c, None)

    # ── Header ────────────────────────────────────────────────────────────────
    _wc(ws, 3, "A", f"员工姓名(Name)：{form.get('employeeName', '')}")
    _wc(ws, 3, "C", f"部门(Apartment)：{form.get('department', '')}")
    _wc(ws, 4, "A", f"常驻地(Permanent Residence)：{form.get('permanentResidence', '')}")
    _wc(ws, 4, "C", f"出差事由(Purpose of Business Trip): {form.get('purpose', '')}")

    # ── Trip info ─────────────────────────────────────────────────────────────
    for i, t in enumerate(trip_info):
        row    = TRIP_R0 + i
        depart = t.get("date", "")
        arrive = t.get("arriveDate", "") or ""
        try:
            d1 = datetime.fromisoformat(depart).strftime("%d-%b-%Y") if depart else ""
            d2 = datetime.fromisoformat(arrive).strftime("%d-%b-%Y") if arrive and arrive != depart else ""
            date_str = f"{d1} - {d2}" if d2 else d1
        except Exception:
            date_str = depart
        _wc(ws, row, "A", date_str)
        _wc(ws, row, "B", f"{t.get('origin','')} - {t.get('destination','')}", wrap=True)
        _wc(ws, row, "C", t.get("vehicle", ""))
        _wc(ws, row, "D", t.get("ticketingMethod", ""))
        # Enforce center alignment on all trip rows (template + inserted)
        for col in ["A", "B", "C", "D"]:
            cell = ws[f"{col}{row}"]
            if not isinstance(cell, MergedCell):
                try:
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=(col == "B"),
                    )
                except Exception:
                    pass

    # ── Transportation ────────────────────────────────────────────────────────
    for i, key in enumerate(all_transport):
        row = trns_r + i
        _wc(ws, row, "A", key, wrap=True)
        _wc(ws, row, "B", intercity_by_route.get(key, 0) or 0)
        _wc(ws, row, "C", urban_by_route.get(key, 0) or 0)
        dcell = ws[f"D{row}"]
        if not isinstance(dcell, MergedCell):
            dcell.value = f"=B{row}+C{row}"

    # ── Accommodation ─────────────────────────────────────────────────────────
    for i, a in enumerate(accommodation):
        row = accom_r + i
        _wc(ws, row, "A", a.get("location", ""))
        _wc(ws, row, "B", a.get("days", 0))
        _wc(ws, row, "C", a.get("amount", 0))

    # ── Allowance ─────────────────────────────────────────────────────────────
    start_date = form.get("allowanceStartDate") or ""
    end_date   = form.get("allowanceEndDate")   or ""
    daily      = form.get("mealAllowanceDailyIDR", 200000)

    if not start_date or not end_date:
        all_dates = sorted(
            [r.get("date", "") for r in receipts] +
            [t.get("date", "") for t in trip_info]
        )
        all_dates = [d for d in all_dates if d]
        if all_dates:
            start_date = start_date or all_dates[0]
            end_date   = end_date   or all_dates[-1]

    ca, cb = ws[f"A{allow_r}"], ws[f"B{allow_r}"]
    if start_date and not isinstance(ca, MergedCell):
        try:
            ca.value = datetime.fromisoformat(start_date)
            ca.number_format = "DD-MMM-YYYY"
        except Exception:
            ca.value = start_date
    if end_date and not isinstance(cb, MergedCell):
        try:
            cb.value = datetime.fromisoformat(end_date)
            cb.number_format = "DD-MMM-YYYY"
        except Exception:
            cb.value = end_date
    if start_date and end_date:
        try:
            days = max((datetime.fromisoformat(end_date) - datetime.fromisoformat(start_date)).days + 1, 1)
            _wc(ws, allow_r, "C", days)
            _wc(ws, allow_r, "D", days * daily)
        except Exception:
            pass

    wb.save(out_path)


# ── Summary page ──────────────────────────────────────────────────────────────

def fill_summary(form: dict, out_path: Path):
    wb = openpyxl.load_workbook(SUMMARY_TEMPLATE)
    ws = wb.active

    if ws.sheet_properties is None:
        ws.sheet_properties = WorksheetProperties()
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.paperSize   = 9
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)

    ws["B2"] = form.get("department", "")
    ws["F2"] = form.get("month", "")

    receipts = sorted(form.get("receipts", []), key=lambda r: r.get("date", ""))

    RECV_R0  = 4; RECV_CAP = 9
    TOTAL_R0 = 13
    PREP_R0  = 16

    extra_recv = max(0, len(receipts) - RECV_CAP)
    if extra_recv > 0:
        _insert_rows_safe(ws, RECV_R0 + RECV_CAP, extra_recv, RECV_R0 + RECV_CAP - 1)

    total_r = TOTAL_R0 + extra_recv
    prep_r  = PREP_R0  + extra_recv

    for row in range(RECV_R0, RECV_R0 + max(RECV_CAP, len(receipts))):
        for c in ["A","B","C","D","E","F","G","H"]:
            _wc(ws, row, c, None)

    category_label = {
        "transportation_intercity": "交通费(城市间)",
        "transportation_urban":     "交通费(市内)",
        "accommodation":            "住宿费",
        "other":                    "其他",
    }

    for i, r in enumerate(receipts):
        row = RECV_R0 + i
        _wc(ws, row, "A", i + 1)
        _wc(ws, row, "B", form.get("employeeName", ""))
        try:
            c_cell = ws[f"C{row}"]
            if not isinstance(c_cell, MergedCell):
                c_cell.value = datetime.fromisoformat(r.get("date", ""))
                c_cell.number_format = "DD-MMM-YYYY"
        except Exception:
            _wc(ws, row, "C", r.get("date", ""))
        _wc(ws, row, "D", category_label.get(r.get("category", ""), r.get("category", "")))
        _wc(ws, row, "E", r.get("currency", "IDR"))
        _wc(ws, row, "F", r.get("amount", 0))

        vendor    = r.get("vendor", "")
        origin    = r.get("origin", "").split(",")[0].strip()
        dest      = r.get("destination", "").split(",")[0].strip()
        reason    = r.get("description", "").strip()
        route_str = f"{origin} - {dest}" if (origin or dest) else ""
        if vendor and route_str:
            note = f"{vendor} ({route_str})"
        elif vendor:
            note = vendor
        elif route_str:
            note = route_str
        else:
            note = ""
        if reason:
            note = f"{note}; {reason}" if note else reason
        _wc(ws, row, "H", note, wrap=True)
        # Expand row height so wrapped remarks text isn't clipped.
        # Estimate ~15pt per line; assume ~60 chars per line at this column width.
        if note:
            lines = max(1, (len(note) + 59) // 60)
            ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 15, lines * 15)

    last_data = RECV_R0 + len(receipts) - 1 if receipts else RECV_R0
    _wc(ws, total_r, "F", f"=SUM(F{RECV_R0}:F{last_data})")
    _wc(ws, prep_r,  "G", "制表人:")   # left blank — filled manually
    _wc(ws, prep_r,  "H", "审核人:")

    wb.save(out_path)


# ── Receipt image → PDF ───────────────────────────────────────────────────────

def image_to_pdf(img_path: Path, pdf_path: Path):
    img = Image.open(img_path).convert("RGB")
    iw, ih = img.size
    aw, ah = A4
    scale = min((aw - 2*cm) / iw, (ah - 2*cm) / ih) * 72 / 96
    rw = min(iw * scale, aw - 2*cm)
    rh = min(ih * scale, ah - 2*cm)
    doc = SimpleDocTemplate(str(pdf_path), pagesize=A4,
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1*cm,  bottomMargin=1*cm)
    doc.build([RLImage(str(img_path), width=rw, height=rh)])


# ── FastAPI ───────────────────────────────────────────────────────────────────

@app.get("/health")
def health():
    return {"status": "ok", "libreoffice": find_soffice() or "not found"}


@app.post("/generate-pdf")
async def generate_pdf(
    form:  str              = Form(...),
    files: List[UploadFile] = File(default=[]),
):
    try:
        form_data = json.loads(form)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid form JSON: {e}")

    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            tmp = Path(tmpdir)

            cover_xlsx   = tmp / "cover.xlsx"
            summary_xlsx = tmp / "summary.xlsx"
            fill_cover(form_data,   cover_xlsx)
            fill_summary(form_data, summary_xlsx)

            cover_pdf   = excel_to_pdf(cover_xlsx,   tmp)
            summary_pdf = excel_to_pdf(summary_xlsx, tmp)

            receipt_pdfs = []
            for i, upload in enumerate(files):
                ext      = Path(upload.filename or "file").suffix.lower()
                raw_path = tmp / f"receipt_{i}{ext}"
                raw_path.write_bytes(await upload.read())
                if ext in (".jpg", ".jpeg", ".png", ".webp"):
                    pdf_path = tmp / f"receipt_{i}.pdf"
                    image_to_pdf(raw_path, pdf_path)
                    receipt_pdfs.append(pdf_path)
                elif ext == ".pdf":
                    receipt_pdfs.append(raw_path)

            merger = PdfMerger()
            merger.append(str(cover_pdf))
            merger.append(str(summary_pdf))
            for rp in receipt_pdfs:
                merger.append(str(rp))

            out_pdf = tmp / "output.pdf"
            merger.write(str(out_pdf))
            merger.close()
            pdf_bytes = out_pdf.read_bytes()

        month = form_data.get("month", "Reimbursement")
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="Reimbursement {month}.pdf"'},
        )

    except Exception as e:
        print(f"[generate-pdf ERROR]\n{traceback.format_exc()}", flush=True)
        raise HTTPException(status_code=500, detail=str(e))
